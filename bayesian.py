from sklearn.feature_extraction import text
import numpy as np
import re, copy
from nltk.corpus import stopwords

## Import excel data of labels
from openpyxl import load_workbook
import numpy as np

## Global variables
heteronyms, vec, fratio, pi0, pi1 = [],[],[],[],[]
numWords = 10 # number of words to be collected from both sides of heteronyms
preNum, postNum = 3, 3 # from left/right of heteronyms
prediction = []

## Search for "word" in "List", returns indices of "word"
def searchList(List, word, start = -1):
    ind = []
    while True:
        try:
            index = List.index(word,start+1)
        except ValueError:
            break
        else:
            ind.append(index)
            start = index
    return ind

## Training: input = training data (.txt), labels(.xlsx)
def bayesianTrain(trainData,trainLabels):
    global heteronyms, vec, fratio, pi0, pi1, numWords, preNum, postNum
    
    datasetFile = trainData + '.txt'
    labelFile = trainLabels + '.xlsx'
    
    wb = load_workbook(filename = labelFile)
    ws = wb['Sheet1']
    
    labels = []
    for row in ws.rows:
        for cell in row:
            labels.append(cell.value)
    
    heteronyms = []
    labelAll = []
    tmpLabel = []
    for i in range(len(labels)):
        if type(labels[i]) == unicode:
            heteronyms.append(str(labels[i]).lower())
            if i != 0:
                labelAll.append(tmpLabel)
            tmpLabel = []
        else:
            if labels[i] is not None:
                tmpLabel.append(int(labels[i]))
            if i == (len(labels) - 1):
                labelAll.append(tmpLabel)

    ## Open text files that contain the word to be trained for

    origText = open(datasetFile).read().split()

    
    ## Look for the word (ind = indices of words)
    ind, wordsAll = [], []
    lenText = len(origText)
    
    for i in range(len(heteronyms)):
        ind.append(searchList(origText,heteronyms[i])) # textlist -> filtered for screening after removing stopwords
        # Collect (numWords) # of nearby words
        words = []
        limit = 0
        for j in ind[i]:
            # Collect upto 300 entries
            limit += 1
            if limit == 301: break
        
            if (j - numWords <= 0) and (j + numWords < lenText):
                words.append([None]*(numWords - j) + origText[0:j+numWords+1])
            elif (j - numWords > 0) and (j + numWords >= lenText):
                words.append(origText[j-numWords:lenText] + [None]*(j+numWords-lenText+1))
            elif (j - numWords <= 0) and (j + numWords >= lenText):
                words.append([None]*(numWords - j) + origText[0:lenText] + [None]*(j+numWords-lenText+1))
            else:
                words.append(origText[j-numWords:j+numWords+1]) 
            
        wordsAll.append(words)
    
    ## 3 arrays!
    preWords, postWords = [], []
    
    for i in range(len(wordsAll)):
        pre, post = [], []
        for j in range(len(wordsAll[i])):
            pre.append(unicode(wordsAll[i][j][numWords-preNum:numWords]))
            post.append(unicode(wordsAll[i][j][numWords+1:numWords+postNum+1]))
            wordsAll[i][j] = unicode(wordsAll[i][j])
        preWords.append(pre)
        postWords.append(post)
    
    ## Vectorizers
    vecAll, vecPre, vecPost = [],[],[]
    fitWords, fitPre, fitPost = [],[],[]
    fitWords0, fitWords1, fitPre0, fitPre1, fitPost0, fitPost1 = [],[],[],[],[],[]
    countW0, countW1, countPre0, countPre1, countPost0, countPost1 = [],[],[],[],[],[]
    W0, W1, Pre0, Pre1, Post0, Post1 = [],[],[],[],[],[]
    fratioW, fratioPre, fratioPost = [],[],[]
    piW0, piW1, piPre0, piPre1, piPost0, piPost1 = [],[],[],[],[],[]
    
    for i in range(len(heteronyms)):
        if len(wordsAll[i]) != 0:
            stop_words = text.ENGLISH_STOP_WORDS.union({heteronyms[i]})
            vecAll.append(text.CountVectorizer(stop_words=stop_words,min_df=2)) # Removes stop words -> words
            vecPre.append(text.CountVectorizer()) # Keeps stop words -> preWords
            vecPost.append(text.CountVectorizer()) # Keeps stop words -> postWords
        
            fitWords.append(vecAll[i].fit_transform(wordsAll[i]).toarray())
            fitPre.append(vecPre[i].fit_transform(preWords[i]).toarray())
            fitPost.append(vecPost[i].fit_transform(postWords[i]).toarray())
        else: # No occurrence of heteronym
            vecAll.append(None)
            vecPre.append(None)
            vecPost.append(None)
            fitWords.append(None)
            fitPre.append(None)
            fitPost.append(None)
        
        # Estimate probability of each word in vocabulary
        k = 0
        fitWords0tmp, fitWords1tmp = [], []
        fitPre0tmp, fitPre1tmp = [], []
        fitPost0tmp, fitPost1tmp = [], []
        for j in labelAll[i]:
            if j == 0: 
                fitWords0tmp.append(fitWords[i][k])
                fitPre0tmp.append(fitPre[i][k])
                fitPost0tmp.append(fitPost[i][k])
            else: 
                fitWords1tmp.append(fitWords[i][k])
                fitPre1tmp.append(fitPre[i][k])
                fitPost1tmp.append(fitPost[i][k])
            k += 1
        fitWords0.append(fitWords0tmp), fitWords1.append(fitWords1tmp)
        fitPre0.append(fitPre0tmp), fitPre1.append(fitPre1tmp)
        fitPost0.append(fitPost0tmp), fitPost1.append(fitPost1tmp)
        
        countW0.append(np.sum(fitWords0[i],axis=0)+1.0), countW1.append(np.sum(fitWords1[i],axis=0)+1.0)
        countPre0.append(np.sum(fitPre0[i],axis=0)+1.0), countPre1.append(np.sum(fitPre1[i],axis=0)+1.0)
        countPost0.append(np.sum(fitPost0[i],axis=0)+1.0), countPost1.append(np.sum(fitPost1[i],axis=0)+1.0)
     
        W0.append(countW0[-1]/np.sum(countW0[-1])), W1.append(countW1[-1]/np.sum(countW1[-1]))
        Pre0.append(countPre0[-1]/np.sum(countPre0[-1])), Pre1.append(countPre1[-1]/np.sum(countPre1[-1]))
        Post0.append(countPost0[-1]/np.sum(countPost0[-1])), Post1.append(countPost1[-1]/np.sum(countPost1[-1]))
        
        # Compute ratio of these probabilities
        fratioW.append(W0[-1]/W1[-1])
        fratioPre.append(Pre0[-1]/Pre1[-1])
        fratioPost.append(Post0[-1]/Post1[-1])
        
        # Compute prior probabilities 
        nW0, nW1 = len(fitWords0[-1]), len(fitWords1[-1])
        nPre0, nPre1= len(fitPre0[-1]), len(fitPre1[-1])
        nPost0, nPost1 = len(fitPost0[-1]), len(fitPost1[-1])
        piW0.append(float(nW0)/(nW0+nW1)), piW1.append(float(nW1)/(nW0+nW1))
        piPre0.append(float(nPre0)/(nPre0+nPre1)), piPre1.append(float(nPre1)/(nPre0+nPre1))
        piPost0.append(float(nPost0)/(nPost0+nPost1)), piPost1.append(float(nPost1)/(nPost0+nPost1))
        
    fratio = [fratioW, fratioPre, fratioPost]
    pi0, pi1 = [piW0, piPre0, piPost0], [piW1, piPre1, piPost1]
    vec = [vecAll, vecPre, vecPost]

def bayesianTest(testingData): 
    global heteronyms, vec, fratio, pi0, pi1, numWords, preNum, postNum
    
    vecAll, vecPre, vecPost = vec[0], vec[1], vec[2]
    fratioW, fratioPre, fratioPost = fratio[0], fratio[1], fratio[2]
    piW0, piPre0, piPost0 = pi0[0], pi0[1], pi0[2]
    piW1, piPre1, piPost1 = pi1[0], pi1[1], pi1[2]
    
#    testFileName = testingData + '.txt'
#    origText = open(testFileName).read()
    origText = testingData
    sentence = re.sub("([^\w']|_)+",' ',origText).lower().split()
    
    ## Locate any heteronym occurrences
    ind, het, words = [], [], []
    preWords, postWords, words, wordsAll = [], [], [], []
    lenText = len(sentence)
    
    numSamples, sampleCt = 40, 0
    for i in range(len(heteronyms)):
        tmpInd = searchList(sentence,heteronyms[i])
        if tmpInd != []:
            if len(tmpInd) >= numSamples:
                tmpInd = tmpInd[0:numSamples]
            
            ind.append(tmpInd)
            het.append(i)
        if sampleCt == numSamples: break
           
    for i in range(len(ind)):
        for j in ind[i]:    
            if (j - numWords <= 0) and (j + numWords < lenText):
                words.append([None]*(numWords - j) + sentence[0:j+numWords+1])
            elif (j - numWords > 0) and (j + numWords >= lenText):
                words.append(sentence[j-numWords:lenText] + [None]*(j+numWords-lenText+1))
            elif (j - numWords <= 0) and (j + numWords >= lenText):
                words.append([None]*(numWords - j) + sentence[0:lenText] + [None]*(j+numWords-lenText+1))
            else:
                words.append(sentence[j-numWords:j+numWords+1])  
            wordsAll.append(words[-1])
            
            # 3 arrays!
            preWords.append(unicode(words[-1][numWords-preNum:numWords]))
            postWords.append(unicode(words[-1][numWords+1:numWords+postNum+1]))
            words[-1] = unicode(words[-1])     
            
    ## Vectorizer
    LR = []
    fitWords, fitPre, fitPost, resultArr = [], [], [], []
    arrInd = 0
    
    for i in range(len(ind)):
        k = het[i]
        for j in ind[i]:
            LRtmp = []
            fitWords.append(vecAll[k].transform([words[arrInd]]).toarray().flatten())
            fitPre.append(vecPre[k].transform([preWords[arrInd]]).toarray().flatten())
            fitPost.append(vecPost[k].transform([postWords[arrInd]]).toarray().flatten())
            LRtmp.append(np.prod(fratioW[k]**fitWords[-1])*piW0[k]/piW1[k])
            LRtmp.append(np.prod(fratioPre[k]**fitPre[-1])*piPre0[k]/piPre1[k])
            LRtmp.append(np.prod(fratioPost[k]**fitPost[-1])*piPost0[k]/piPost1[k])
            LR.append(LRtmp)
            
            result, ct = 0,0
            ctLabels = ["All", "Pre", "Post"]
            
            for lr in LRtmp:
                if lr == 0.57142857142857151: # All elements = "None"
                    result += 0
                elif lr > 2:
                    result += 2
                elif (lr <= 2) & (lr > 1):
                    result += 1
                elif (lr <= 1) & (lr > 0.5):
                    result -= 1
                else:
                    result -= 2
                
            if result >= 0:
                result = 0
#                print heteronyms[k] +"\t["+dic[k][0]+"]"
            else:
                result = 1
#                print heteronyms[k] +"\t["+dic[k][1]+"]"
            
            resultArr.append(result)
            arrInd += 1

    return resultArr, ind

def bayesianAccuracy(testingLabels):
    global prediction
        
    ## Load labels for test data
    testingLabelFile = testingLabels + '.xlsx'
    wb = load_workbook(filename = testingLabelFile)
    ws = wb['Sheet1']
    
    OANC_labels = []
    for row in ws.rows:
        for cell in row:
            OANC_labels.append(cell.value)
            
    hetTest = []
    labelTest = []
    tmpLabel = []
    for i in range(len(OANC_labels)):
        if type(OANC_labels[i]) == unicode:
            hetTest.append(str(OANC_labels[i]).lower())
            if i != 0:
                labelTest.append(tmpLabel)
            tmpLabel = []
        else:
            if OANC_labels[i] is not None:
                tmpLabel.append(int(OANC_labels[i]))
            if i == (len(OANC_labels) - 1):
                labelTest.append(tmpLabel)        
    del wb, ws, OANC_labels, tmpLabel
    
    ## Compare classification result with test data labels
    percent = []
    correct, b = 0, 0
    for a in [0,1,3,4]:
        correct = len([i for i, j in zip(prediction[a*40:a*40+40], labelTest[b]) if i == j])
        percent.append(correct/40.0*100)
        b += 1
    
    return percent


#testData = 'OANC'
#testLabels = 'OANC_labels'

def Run(testData):
    trainData = 'textDataAll'
    trainLabels = 'labels'
    bayesianTrain(trainData, trainLabels)
    prediction, ind = bayesianTest(testData)
    return prediction, ind
    #percent = bayesianAccuracy(testLabels)