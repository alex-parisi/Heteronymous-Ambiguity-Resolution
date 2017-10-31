import os
import copy
from sklearn import tree
from sklearn.feature_extraction import text
import numpy as np
from openpyxl import load_workbook

# disable annoying warnings
import warnings
warnings.filterwarnings("ignore")
    
# actual beginning     
numWords = 10
clf = []
clfPre = []
clfPost = []
vectorizer = []
vectorizerPre = []
vectorizerPost = []
trainData = []
heteronyms = []
preNum, postNum = 3, 3 # from left/right of heteronyms


# Search for a specific heteronym
def searchList(List, word, start = -1):
    ind = []
    while True:
        try:
            index = List.index(word,start+1)
        except ValueError:
            break
        else:
            ind.append(index)
            start = index
    return ind

# Train feature extraction
def trainVectorizer():
    
    global numWords
    global vectorizer
    global vectorizerPre
    global vectorizerPost
    global trainData
    global heteronyms
    
    wordsAll = []
    ind = []
    features = []
    featuresPre = []
    featuresPost = []
    lenText = len(trainData)
    
    # Look for the word (ind = indices of words)
    for i in range(len(heteronyms)):
        limit = 0
        ind.append(searchList(trainData,heteronyms[i])) # textlist -> filtered for screening after removing stopwords
        # Collect (numWords) # of nearby words
        words = []
        for j in ind[i]:
            limit = limit + 1
            if (j - numWords <= 0) and (j + numWords < lenText):
                words.append([None]*(numWords - j) + trainData[0:j+numWords+1])
            elif (j - numWords > 0) and (j + numWords >= lenText):
                words.append(trainData[j-numWords:lenText] + [None]*(j+numWords-lenText+1))
            elif (j - numWords <= 0) and (j + numWords >= lenText):
                words.append([None]*(numWords - j) + trainData[0:lenText] + [None]*(j+numWords-lenText+1))
            else:
                words.append(trainData[j-numWords:j+numWords+1])
            if limit >= 300:
                break
        wordsAll.append(words)

    ## 3 arrays!
    preWords, postWords = [], []
    
    for i in range(len(wordsAll)):
        pre, post = [], []
        for j in range(len(wordsAll[i])):
            pre.append(unicode(wordsAll[i][j][numWords-preNum:numWords]))
            post.append(unicode(wordsAll[i][j][numWords+1:numWords+postNum+1]))
            wordsAll[i][j] = unicode(wordsAll[i][j])
        preWords.append(pre)
        postWords.append(post)


    for i in range(len(wordsAll)):
        for j in range(len(wordsAll[i])):
            wordsAll[i][j] = unicode(wordsAll[i][j])



    for i in range(len(heteronyms)):
        if len(wordsAll[i]) != 0:
            #stop_words = text.ENGLISH_STOP_WORDS.union(heteronyms[i])
            vectorizer.append(text.CountVectorizer(min_df=2)) # Removes stop words -> words
            vectorizerPre.append(text.CountVectorizer(min_df=2))
            vectorizerPost.append(text.CountVectorizer(min_df=2))
            features.append(vectorizer[i].fit_transform(wordsAll[i]).toarray())
            featuresPre.append(vectorizerPre[i].fit_transform(preWords[i]).toarray())
            featuresPost.append(vectorizerPost[i].fit_transform(postWords[i]).toarray())
        else: # No occurrence of heteronym
            vectorizer.append([])
            vectorizerPre.append([])
            vectorizerPost.append([])
            features.append([])
            featuresPre.append([])
            featuresPost.append([])

    return features, featuresPre, featuresPost

# Extract features from one single text file
def extractFeatures(origText):
    
    global numWords
    global vectorizer
    global vectorizerPre
    global vectorizerPost
    global heteronyms
    
    words = []
    ind = []
    features = []
    featuresPre = []
    featuresPost = []
    wordsAll = []
    if isinstance(origText, (str, unicode)) == True:
        origText = origText.split()
    lenText = len(origText)
    
    # Look for the word (ind = indices of words)
    for i in range(len(heteronyms)):
        limit = 0
        ind.append(searchList(origText,heteronyms[i])) # textlist -> filtered for screening after removing stopwords
        # Collect (numWords) # of nearby words
        words = []
        for j in ind[i]:
            limit = limit + 1
            if (j - numWords <= 0) and (j + numWords < lenText):
                words.append([None]*(numWords - j) + origText[0:j+numWords+1])
            elif (j - numWords > 0) and (j + numWords >= lenText):
                words.append(origText[j-numWords:lenText] + [None]*(j+numWords-lenText+1))
            elif (j - numWords <= 0) and (j + numWords >= lenText):
                words.append([None]*(numWords - j) + origText[0:lenText] + [None]*(j+numWords-lenText+1))
            else:
                words.append(origText[j-numWords:j+numWords+1])
            if limit >= 40:
                break
        wordsAll.append(words)

    ## 3 arrays!
    preWords, postWords = [], []

    for i in range(len(wordsAll)):
        pre, post = [], []
        for j in range(len(wordsAll[i])):
            pre.append(unicode(wordsAll[i][j][numWords-preNum:numWords]))
            post.append(unicode(wordsAll[i][j][numWords+1:numWords+postNum+1]))
            wordsAll[i][j] = unicode(wordsAll[i][j])
        preWords.append(pre)
        postWords.append(post)

    for i in range(len(wordsAll)):
        for j in range(len(wordsAll[i])):
            wordsAll[i][j] = unicode(wordsAll[i][j])
        if len(wordsAll[i]) != 0:
            features.append(vectorizer[i].transform(wordsAll[i]).toarray())
            featuresPre.append(vectorizerPre[i].transform(preWords[i]).toarray())
            featuresPost.append(vectorizerPost[i].transform(postWords[i]).toarray())
        
        else: # No occurrence of heteronym
            
            features.append([])
            featuresPre.append([])
            featuresPost.append([])

    return ind, features, featuresPre, featuresPost

def trainTree():
    
    global clf
    global clfPre
    global clfPost
    global vectorizer
    global vectorizerPre
    global vectorizerPost
    global trainData
    
    fileDir = os.path.dirname(os.path.realpath('__file__'))
    trainPath = os.path.join(fileDir,'../data/textDataAll.txt')
    trainPath = os.path.abspath(os.path.realpath(trainPath))
    trainData = open(trainPath).read().split()
    
    labPath = os.path.join(fileDir, '../data/labels.xlsx')
    wb = load_workbook(labPath)
    ws = wb['Sheet1']
    
    labels = []
    for row in ws.rows:
        for cell in row:
            labels.append(cell.value)
    tmpLabel = []
    labelAll = []
    for i in range(len(labels)):
        if type(labels[i]) == unicode:
            heteronyms.append(str(labels[i]).lower())
            if i != 0:
                labelAll.append(tmpLabel)
            tmpLabel = []
        else:
            if labels[i] is not None:
                tmpLabel.append(int(labels[i]))
            if i == (len(labels) - 1):
                labelAll.append(tmpLabel)

    features, featuresPre, featuresPost = trainVectorizer()
    
    for i in range(len(heteronyms)):
        clf.append(tree.DecisionTreeClassifier())
        clfPre.append(tree.DecisionTreeClassifier())
        clfPost.append(tree.DecisionTreeClassifier())
        clf[i] = clf[i].fit(features[i],labelAll[i])
        clfPre[i] = clfPre[i].fit(featuresPre[i],labelAll[i])
        clfPost[i] = clfPost[i].fit(featuresPost[i],labelAll[i])


def testTree(testData):
    
    global heteronyms
    global clf
    global clfPre
    global clfPost
    
    features, featuresPre, featuresPost = [],[],[]
    ind, features, featuresPre, featuresPost = extractFeatures(testData)
    prediction, predictionPre, predictionPost = [],[],[]
        

    for i in range(len(heteronyms)):
        if len(features[i]) != 0:
            prediction1 = []
            predictionPre1 = []
            predictionPost1 = []
            p, pPre, pPost, p1, p2, p3 = [],[],[],[],[],[]
            for j in range(len(features[i])):
                prediction1.append(clf[i].predict(features[i][j]))
                predictionPre1.append(clfPre[i].predict(featuresPre[i][j]))
                predictionPost1.append(clfPost[i].predict(featuresPost[i][j]))
                p.append(clf[i].predict_proba(features[i][j]))
                pPre.append(clfPre[i].predict_proba(featuresPre[i][j]))
                pPost.append(clfPost[i].predict_proba(featuresPost[i][j]))
            
            prediction.append(prediction1)
            predictionPre.append(predictionPre1)
            predictionPost.append(predictionPost1)
            p1.append(p)
            p2.append(pPre)
            p3.append(pPost)

    return ind, prediction, predictionPre, predictionPost, p1, p2, p3

def accuracy(prediction): 
    
    testDir = os.path.dirname(os.path.realpath('__file__'))
    labPath = os.path.join(testDir, '../data/testLabels.xlsx')
    wb = load_workbook(labPath)
    ws = wb['Sheet1']
    
    OANC_labels = []
    for row in ws.rows:
        for cell in row:
            OANC_labels.append(cell.value)
            
    hetTest = []
    labelTest = []
    tmpLabel = []
    for i in range(len(OANC_labels)):
        if type(OANC_labels[i]) == unicode:
            hetTest.append(str(OANC_labels[i]).lower())
            if i != 0:
                labelTest.append(tmpLabel)
            tmpLabel = []
        else:
            if OANC_labels[i] is not None:
                tmpLabel.append(int(OANC_labels[i]))
            if i == (len(OANC_labels) - 1):
                labelTest.append(tmpLabel)        
    del wb, ws, OANC_labels, tmpLabel
    
    ## Compare classification result with test data labels
    percent = []
    correct, b = 0, 0
    for a in [0,1,3,4]:
        correct = len([i for i, j in zip(prediction[a], labelTest[b]) if i == j])
        percent.append(correct/40.0*100.0)
        b += 1
    print percent

    

#   MAIN
def Run(testData):
    trainTree()
    ind,prediction, predictionPre, predictionPost, p1, p2, p3 = testTree(testData)
    
    finalPrediction = copy.deepcopy(prediction)
    for i in range(len(prediction)):
        for j in range(len(prediction[i])):
            finalPrediction[i][j] = np.round((prediction[i][j]+predictionPre[i][j] + predictionPost[i][j])/3.0)
    return finalPrediction, ind



# TESTING:
#testDir = os.path.dirname(os.path.realpath('__file__'))
#testPath = os.path.join(testDir,'../data/OANC.txt')
#testPath = os.path.abspath(os.path.realpath(testPath))
#testData = open(testPath).read()
##testData = "I want you to tell me what to cook, close the door and spend the afternoon just hanging out wiht me"
#
#trainTree()
#ind,prediction, predictionPre, predictionPost, p1, p2, p3 = testTree(testData)
#
#finalPrediction = copy.deepcopy(prediction)
#for i in range(len(prediction)):
#    for j in range(len(prediction[i])):
#        finalPrediction[i][j] = np.round((prediction[i][j]+predictionPre[i][j] + predictionPost[i][j])/3.0)
#print(finalPrediction)
#
##percent = accuracy(finalPrediction)
